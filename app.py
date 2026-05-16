
Action: file_editor create /app/output/app.py --file-text "\"\"\"
═══════════════════════════════════════════════════════════════
  Billing System Pro - Flask Backend (PostgreSQL)
  HTML/CSS/JS are now in separate files:
      templates/index.html
      static/style.css
      static/script.js
═══════════════════════════════════════════════════════════════
\"\"\"
import os, hashlib, io, calendar
from datetime import datetime, date
from functools import wraps
from flask import Flask, request, jsonify, session, render_template, send_file
from flask_cors import CORS

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = os.environ.get(\"SECRET_KEY\", \"billpro_secret_key_2024\")
CORS(app, supports_credentials=True)

# ══════════════════════════════════════════════════════════
# DATABASE SETUP
# ══════════════════════════════════════════════════════════
DATABASE_URL = os.environ.get(\"DATABASE_URL\", \"\")
if DATABASE_URL.startswith(\"postgres://\"):
    DATABASE_URL = DATABASE_URL.replace(\"postgres://\", \"postgresql://\", 1)

DB_READY = False
DB_ERROR = \"\"

try:
    import psycopg2
    import psycopg2.extras
    if not DATABASE_URL:
        raise Exception(\"DATABASE_URL environment variable is not set.\")
    test_conn = psycopg2.connect(DATABASE_URL)
    test_conn.close()
    DB_READY = True
except ImportError:
    DB_ERROR = \"psycopg2 not installed. Check requirements.txt has psycopg2-binary.\"
except Exception as e:
    DB_ERROR = str(e)


def get_db():
    return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)


def qone(sql, params=()):
    conn = get_db(); cur = conn.cursor()
    cur.execute(sql, params)
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None


def qall(sql, params=()):
    conn = get_db(); cur = conn.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


def run(sql, params=()):
    conn = get_db(); cur = conn.cursor()
    cur.execute(sql, params)
    conn.commit(); conn.close()


def hpw(pw):
    return hashlib.sha256(pw.strip().encode()).hexdigest()


# ══════════════════════════════════════════════════════════
# INIT TABLES
# ══════════════════════════════════════════════════════════
def init_db():
    conn = get_db(); cur = conn.cursor()

    cur.execute(\"\"\"CREATE TABLE IF NOT EXISTS products (
        code TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        price NUMERIC NOT NULL DEFAULT 0,
        stock INTEGER NOT NULL DEFAULT 0)\"\"\")

    cur.execute(\"\"\"CREATE TABLE IF NOT EXISTS workers (
        number TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        created TIMESTAMP DEFAULT CURRENT_TIMESTAMP)\"\"\")

    cur.execute(\"\"\"CREATE TABLE IF NOT EXISTS bills (
        id SERIAL PRIMARY KEY,
        bill_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        customer_name TEXT NOT NULL,
        customer_phone TEXT NOT NULL,
        customer_email TEXT DEFAULT '',
        customer_addr TEXT DEFAULT '',
        worker_number TEXT DEFAULT '',
        worker_name TEXT DEFAULT '',
        total_amount NUMERIC DEFAULT 0,
        total_pieces INTEGER DEFAULT 0)\"\"\")

    cur.execute(\"\"\"CREATE TABLE IF NOT EXISTS bill_items (
        id SERIAL PRIMARY KEY,
        bill_id INTEGER NOT NULL REFERENCES bills(id),
        product_code TEXT,
        product_name TEXT,
        price NUMERIC,
        quantity INTEGER,
        subtotal NUMERIC)\"\"\")

    cur.execute(\"\"\"CREATE TABLE IF NOT EXISTS adjustments (
        id SERIAL PRIMARY KEY,
        worker_number TEXT NOT NULL,
        pieces INTEGER NOT NULL,
        note TEXT DEFAULT '',
        created TIMESTAMP DEFAULT CURRENT_TIMESTAMP)\"\"\")

    # ── SUPERVISOR ────────────────────────────────────────
    cur.execute(\"\"\"CREATE TABLE IF NOT EXISTS supervisor (
        id SERIAL PRIMARY KEY,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        created TIMESTAMP DEFAULT CURRENT_TIMESTAMP)\"\"\")

    # ── MANAGER (new) ─────────────────────────────────────
    cur.execute(\"\"\"CREATE TABLE IF NOT EXISTS manager (
        id SERIAL PRIMARY KEY,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL)\"\"\")

    # ── WORKER ASSIGNMENT (new) ───────────────────────────
    cur.execute(\"\"\"CREATE TABLE IF NOT EXISTS supervisor_workers (
        id SERIAL PRIMARY KEY,
        supervisor_id INTEGER NOT NULL REFERENCES supervisor(id) ON DELETE CASCADE,
        worker_number TEXT NOT NULL,
        UNIQUE(worker_number))\"\"\")

    # ── ATTENDANCE (new) ──────────────────────────────────
    cur.execute(\"\"\"CREATE TABLE IF NOT EXISTS attendance (
        id SERIAL PRIMARY KEY,
        worker_number TEXT NOT NULL,
        att_date DATE NOT NULL,
        status TEXT NOT NULL,
        marked_by INTEGER,
        marked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(worker_number, att_date))\"\"\")

    # ── seed default supervisor ───────────────────────────
    cur.execute(\"SELECT COUNT(*) AS n FROM supervisor\")
    if cur.fetchone()['n'] == 0:
        cur.execute(\"INSERT INTO supervisor(username,password) VALUES(%s,%s)\",
                    ('admin', hpw('admin123')))

    # ── seed default manager ──────────────────────────────
    cur.execute(\"SELECT COUNT(*) AS n FROM manager\")
    if cur.fetchone()['n'] == 0:
        cur.execute(\"INSERT INTO manager(username,password) VALUES(%s,%s)\",
                    ('manager', hpw('manager123')))

    conn.commit(); conn.close()


if DB_READY:
    try:
        init_db()
        print(\"✅ Database tables ready.\")
    except Exception as e:
        DB_READY = False
        DB_ERROR = \"Tables init failed: \" + str(e)
        print(\"❌ \" + DB_ERROR)
else:
    print(\"❌ DB not ready: \" + DB_ERROR)


# ══════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════
def need_sup(fn):
    @wraps(fn)
    def wrap(*a, **kw):
        if not session.get(\"is_sup\") and not session.get(\"is_mgr\"):
            return jsonify({\"error\": \"Supervisor login required\"}), 401
        return fn(*a, **kw)
    return wrap


def need_mgr(fn):
    @wraps(fn)
    def wrap(*a, **kw):
        if not session.get(\"is_mgr\"):
            return jsonify({\"error\": \"Manager login required\"}), 401
        return fn(*a, **kw)
    return wrap


def jok(**kwargs):
    kwargs['ok'] = True
    return jsonify(kwargs)


def jerr(msg, code=400):
    return jsonify({'error': msg}), code


def jdata():
    d = request.get_json(force=True, silent=True)
    return d if isinstance(d, dict) else {}


def db_check():
    if not DB_READY:
        return jerr(\"Database not connected. Error: \" + DB_ERROR, 503)
    return None


# ══════════════════════════════════════════════════════════
# FRONTEND
# ══════════════════════════════════════════════════════════
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/db-status', methods=['GET'])
def db_status():
    if not DB_READY:
        return jsonify({'ok': False, 'error': DB_ERROR})
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════
# SUPERVISOR AUTH
# ══════════════════════════════════════════════════════════
@app.route('/api/supervisor/login', methods=['POST'])
def sup_login():
    err = db_check()
    if err: return err
    d = jdata()
    u = str(d.get('username', '') or '').strip()
    p = str(d.get('password', '') or '').strip()
    if not u or not p:
        return jerr('Username and password required')
    row = qone(\"SELECT * FROM supervisor WHERE username=%s AND password=%s\",
               (u, hpw(p)))
    if not row:
        return jerr('Wrong username or password', 401)
    session.clear()
    session['is_sup'] = True
    session['sup_id'] = row['id']
    session['sup_u'] = u
    return jok(username=u, role='supervisor')


@app.route('/api/supervisor/register', methods=['POST'])
def sup_register():
    err = db_check()
    if err: return err
    d = jdata()
    u = str(d.get('username', '') or '').strip()
    p = str(d.get('password', '') or '').strip()
    if not u or len(u) < 3:
        return jerr('Username must be at least 3 characters')
    if not p or len(p) < 4:
        return jerr('Password must be at least 4 characters')
    if qone(\"SELECT id FROM supervisor WHERE username=%s\", (u,)):
        return jerr(f'Username \"{u}\" already exists')
    run(\"INSERT INTO supervisor(username,password) VALUES(%s,%s)\", (u, hpw(p)))
    return jok(message=f'Supervisor \"{u}\" created. You can now log in.')


@app.route('/api/supervisor/logout', methods=['POST'])
def sup_logout():
    session.clear()
    return jok()


@app.route('/api/supervisor/status', methods=['GET'])
def sup_status():
    return jsonify({
        'logged_in': bool(session.get('is_sup') or session.get('is_mgr')),
        'is_supervisor': bool(session.get('is_sup')),
        'is_manager': bool(session.get('is_mgr')),
        'username': session.get('sup_u') or session.get('mgr_u') or '',
        'sup_id': session.get('sup_id')
    })


# ══════════════════════════════════════════════════════════
# MANAGER AUTH
# ══════════════════════════════════════════════════════════
@app.route('/api/manager/login', methods=['POST'])
def mgr_login():
    err = db_check()
    if err: return err
    d = jdata()
    u = str(d.get('username', '') or '').strip()
    p = str(d.get('password', '') or '').strip()
    if not u or not p:
        return jerr('Username and password required')
    row = qone(\"SELECT * FROM manager WHERE username=%s AND password=%s\",
               (u, hpw(p)))
    if not row:
        return jerr('Wrong manager credentials', 401)
    session.clear()
    session['is_mgr'] = True
    session['mgr_id'] = row['id']
    session['mgr_u'] = u
    return jok(username=u, role='manager')


@app.route('/api/manager/logout', methods=['POST'])
def mgr_logout():
    session.clear()
    return jok()


# ══════════════════════════════════════════════════════════
# PRODUCTS
# ══════════════════════════════════════════════════════════
@app.route('/api/products', methods=['GET'])
def get_products():
    err = db_check()
    if err: return err
    return jsonify(qall(\"SELECT * FROM products ORDER BY code\"))


@app.route('/api/products/<code>', methods=['GET'])
def get_product(code):
    err = db_check()
    if err: return err
    p = qone(\"SELECT * FROM products WHERE code=%s\", (code.strip(),))
    if p: return jsonify(p)
    return jerr('Product not found', 404)


@app.route('/api/products', methods=['POST'])
def add_product():
    err = db_check()
    if err: return err
    d = jdata()
    code = str(d.get('code', '') or '').strip()
    name = str(d.get('name', '') or '').strip()
    price = d.get('price'); stock = d.get('stock')
    if not code or len(code) != 3 or not code.isdigit():
        return jerr('Product code must be exactly 3 digits')
    if not name: return jerr('Product name is required')
    try: price = float(price)
    except: return jerr('Enter a valid price number')
    try: stock = int(stock)
    except: return jerr('Enter a valid stock quantity')
    if stock < 0: return jerr('Stock cannot be negative')
    if qone(\"SELECT code FROM products WHERE code=%s\", (code,)):
        return jerr(f'Product code {code} already exists')
    run(\"INSERT INTO products(code,name,price,stock) VALUES(%s,%s,%s,%s)\",
        (code, name, price, stock))
    return jok(message=f'Product \"{name}\" added')


@app.route('/api/products/<code>', methods=['DELETE'])
def del_product(code):
    err = db_check()
    if err: return err
    run(\"DELETE FROM products WHERE code=%s\", (code,))
    return jok()


# ══════════════════════════════════════════════════════════
# WORKERS
# ══════════════════════════════════════════════════════════
@app.route('/api/workers', methods=['GET'])
def get_workers():
    err = db_check()
    if err: return err
    return jsonify(qall(\"SELECT * FROM workers ORDER BY number\"))


@app.route('/api/workers/<number>', methods=['GET'])
def get_worker(number):
    err = db_check()
    if err: return err
    w = qone(\"SELECT * FROM workers WHERE number=%s\", (number.strip(),))
    if w: return jsonify(w)
    return jerr('Worker not found', 404)


@app.route('/api/workers', methods=['POST'])
def add_worker():
    err = db_check()
    if err: return err
    d = jdata()
    num = str(d.get('number', '') or '').strip()
    name = str(d.get('name', '') or '').strip()
    if not num: return jerr('Worker number is required')
    if not name: return jerr('Worker name is required')
    if qone(\"SELECT number FROM workers WHERE number=%s\", (num,)):
        return jerr(f'Worker {num} already exists')
    run(\"INSERT INTO workers(number,name) VALUES(%s,%s)\", (num, name))
    return jok(message=f'Worker \"{name}\" added')


@app.route('/api/workers/<number>', methods=['DELETE'])
def del_worker(number):
    err = db_check()
    if err: return err
    run(\"DELETE FROM workers WHERE number=%s\", (number,))
    run(\"DELETE FROM supervisor_workers WHERE worker_number=%s\", (number,))
    run(\"DELETE FROM attendance WHERE worker_number=%s\", (number,))
    return jok()


# ══════════════════════════════════════════════════════════
# BILLS
# ══════════════════════════════════════════════════════════
@app.route('/api/bills/next-id', methods=['GET'])
def next_id():
    err = db_check()
    if err: return err
    r = qone(\"SELECT COALESCE(MAX(id),0) AS m FROM bills\")
    return jsonify({'next_id': r['m'] + 1})


@app.route('/api/bills', methods=['POST'])
def create_bill():
    err = db_check()
    if err: return err
    d = jdata()
    cname = str(d.get('customer_name', '') or '').strip()
    cphone = str(d.get('customer_phone', '') or '').strip()
    cemail = str(d.get('customer_email', '') or '').strip()
    caddr = str(d.get('customer_addr', '') or '').strip()
    wnum = str(d.get('worker_number', '') or '').strip()
    wname = str(d.get('worker_name', '') or '').strip()
    items = d.get('items') or []
    if not cname: return jerr('Customer name is required')
    if not cphone or len(cphone) != 10 or not cphone.isdigit():
        return jerr('Phone must be exactly 10 digits')
    if not items: return jerr('Add at least one item')
    if wnum and not qone(\"SELECT number FROM workers WHERE number=%s\", (wnum,)):
        return jerr(f'Worker {wnum} not found. Add worker first.')
    conn = get_db(); cur = conn.cursor()
    total_a = 0; total_p = 0; validated = []
    for item in items:
        code = str(item.get('code', '') or '').strip()
        qty = int(item.get('quantity', 1) or 1)
        cur.execute(\"SELECT * FROM products WHERE code=%s\", (code,))
        p = cur.fetchone()
        if not p:
            conn.close(); return jerr(f'Product code {code} not found')
        if p['stock'] < qty:
            conn.close(); return jerr(f'Not enough stock for {p[\"name\"]} (available: {p[\"stock\"]})')
        sub = float(p['price']) * qty
        total_a += sub; total_p += qty
        validated.append({'code': code, 'name': p['name'],
                          'price': float(p['price']), 'qty': qty, 'sub': sub})
    cur.execute(\"\"\"INSERT INTO bills(customer_name,customer_phone,customer_email,customer_addr,
                   worker_number,worker_name,total_amount,total_pieces)
                   VALUES(%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id\"\"\",
                (cname, cphone, cemail, caddr, wnum, wname, total_a, total_p))
    bid = cur.fetchone()['id']
    for it in validated:
        cur.execute(\"\"\"INSERT INTO bill_items(bill_id,product_code,product_name,price,quantity,subtotal)
                       VALUES(%s,%s,%s,%s,%s,%s)\"\"\",
                    (bid, it['code'], it['name'], it['price'], it['qty'], it['sub']))
        cur.execute(\"UPDATE products SET stock=stock-%s WHERE code=%s\", (it['qty'], it['code']))
    conn.commit(); conn.close()
    return jok(bill_id=bid, total=total_a, pieces=total_p)


@app.route('/api/customers/lookup', methods=['GET'])
def lookup_cust():
    err = db_check()
    if err: return err
    ph = (request.args.get('phone') or '').strip()
    if not ph: return jerr('Phone required')
    b = qone(\"SELECT * FROM bills WHERE customer_phone=%s ORDER BY id DESC LIMIT 1\", (ph,))
    if not b: return jerr(f'No bills found for {ph}', 404)
    b['items'] = qall(\"SELECT * FROM bill_items WHERE bill_id=%s\", (b['id'],))
    b['total_count'] = qone(\"SELECT COUNT(*) AS c FROM bills WHERE customer_phone=%s\", (ph,))['c']
    return jsonify(b)


# ══════════════════════════════════════════════════════════
# INCENTIVES
# ══════════════════════════════════════════════════════════
@app.route('/api/incentives', methods=['GET'])
def get_incentives():
    err = db_check()
    if err: return err
    workers = qall(\"SELECT * FROM workers ORDER BY number\")
    out = []
    for w in workers:
        bd = qone(\"SELECT COUNT(*) AS cnt, COALESCE(SUM(total_pieces),0) AS pcs FROM bills WHERE worker_number=%s\",
                  (w['number'],))
        adj = qone(\"SELECT COALESCE(SUM(pieces),0) AS tot FROM adjustments WHERE worker_number=%s\",
                   (w['number'],))
        p = int(bd['pcs'] or 0) + int(adj['tot'] or 0)
        out.append({'number': w['number'], 'name': w['name'],
                    'pieces': p, 'bills': bd['cnt'], 'incentive': p})
    return jsonify(out)


@app.route('/api/incentives/adjust', methods=['POST'])
@need_sup
def adj_inc():
    err = db_check()
    if err: return err
    d = jdata()
    wnum = str(d.get('worker_number', '') or '').strip()
    note = str(d.get('note', '') or '').strip()
    try: adj = int(d.get('adjustment', 0))
    except: return jerr('Invalid adjustment')
    if not wnum: return jerr('Worker number required')
    if not qone(\"SELECT number FROM workers WHERE number=%s\", (wnum,)):
        return jerr(f'Worker {wnum} not found')
    if adj == 0: return jerr('Adjustment cannot be zero')
    run(\"INSERT INTO adjustments(worker_number,pieces,note) VALUES(%s,%s,%s)\",
        (wnum, adj, note))
    return jok(message=f'Adjusted {adj:+d} pieces for worker {wnum}')


@app.route('/api/incentives/clear', methods=['POST'])
@need_sup
def clr_inc():
    err = db_check()
    if err: return err
    conn = get_db(); cur = conn.cursor()
    cur.execute(\"UPDATE bills SET worker_number='',worker_name=''\")
    cur.execute(\"DELETE FROM adjustments\")
    conn.commit(); conn.close()
    return jok(message='All incentives cleared for new month')


# ══════════════════════════════════════════════════════════
# REPORTS (overall)
# ══════════════════════════════════════════════════════════
@app.route('/api/reports', methods=['GET'])
def get_reports():
    err = db_check()
    if err: return err
    sales = qone(\"SELECT COALESCE(SUM(total_amount),0) AS v FROM bills\")['v']
    nbills = qone(\"SELECT COUNT(*) AS v FROM bills\")['v']
    ncusts = qone(\"SELECT COUNT(DISTINCT customer_phone) AS v FROM bills\")['v']
    rows = qall(\"\"\"SELECT COALESCE(SUM(b.total_pieces),0) AS inc FROM workers w
                   LEFT JOIN bills b ON b.worker_number=w.number GROUP BY w.number\"\"\")
    tinc = sum(int(r['inc'] or 0) for r in rows)
    recent = qall(\"\"\"SELECT id,bill_date,customer_name,customer_phone,total_amount,worker_number
                     FROM bills ORDER BY id DESC LIMIT 15\"\"\")
    top = qall(\"\"\"SELECT product_name,SUM(quantity) AS units,SUM(subtotal) AS revenue
                  FROM bill_items GROUP BY product_name ORDER BY units DESC LIMIT 10\"\"\")
    return jsonify({'total_sales': float(sales), 'total_bills': nbills,
                    'total_customers': ncusts, 'total_incentives': tinc,
                    'recent_bills': recent, 'top_products': top})


# ══════════════════════════════════════════════════════════
# ATTENDANCE  -  Supervisor / Manager
# ══════════════════════════════════════════════════════════
def workers_for_session():
    \"\"\"Returns the list of workers the current session can see/edit.\"\"\"
    if session.get('is_mgr'):
        return qall(\"SELECT * FROM workers ORDER BY number\")
    if session.get('is_sup'):
        sid = session.get('sup_id')
        return qall(\"\"\"SELECT w.* FROM workers w
                       JOIN supervisor_workers sw ON sw.worker_number=w.number
                       WHERE sw.supervisor_id=%s ORDER BY w.number\"\"\", (sid,))
    return []


@app.route('/api/attendance/my-workers', methods=['GET'])
@need_sup
def my_workers():
    err = db_check()
    if err: return err
    return jsonify(workers_for_session())


@app.route('/api/attendance', methods=['GET'])
@need_sup
def get_attendance():
    \"\"\"For a given date, return assigned workers with their status.\"\"\"
    err = db_check()
    if err: return err
    qdate = (request.args.get('date') or date.today().isoformat()).strip()
    try:
        datetime.strptime(qdate, '%Y-%m-%d')
    except:
        return jerr('Invalid date (use YYYY-MM-DD)')

    ws = workers_for_session()
    if not ws:
        return jsonify({'date': qdate, 'rows': []})

    nums = [w['number'] for w in ws]
    placeholders = ','.join(['%s'] * len(nums))
    recs = qall(f\"\"\"SELECT worker_number,status FROM attendance
                    WHERE att_date=%s AND worker_number IN ({placeholders})\"\"\",
                tuple([qdate] + nums))
    status_map = {r['worker_number']: r['status'] for r in recs}

    rows = []
    for w in ws:
        # If supervisor, also show which supervisor 'owns' this worker (manager view)
        sup_name = ''
        if session.get('is_mgr'):
            sw = qone(\"\"\"SELECT s.username FROM supervisor_workers sw
                         JOIN supervisor s ON s.id=sw.supervisor_id
                         WHERE sw.worker_number=%s\"\"\", (w['number'],))
            sup_name = sw['username'] if sw else '— unassigned —'
        rows.append({
            'number': w['number'],
            'name': w['name'],
            'status': status_map.get(w['number'], ''),
            'supervisor': sup_name
        })
    return jsonify({'date': qdate, 'rows': rows})


@app.route('/api/attendance/mark', methods=['POST'])
@need_sup
def mark_attendance():
    err = db_check()
    if err: return err
    d = jdata()
    wnum = str(d.get('worker_number', '') or '').strip()
    qdate = str(d.get('date', '') or date.today().isoformat()).strip()
    status = str(d.get('status', '') or '').strip().upper()

    if status not in ('P', 'A', 'H', 'L'):
        return jerr('Status must be P / A / H / L')
    if not wnum:
        return jerr('Worker number required')
    try:
        datetime.strptime(qdate, '%Y-%m-%d')
    except:
        return jerr('Invalid date')

    # Permission: supervisor can mark only their own workers
    if session.get('is_sup') and not session.get('is_mgr'):
        owns = qone(\"\"\"SELECT id FROM supervisor_workers
                       WHERE supervisor_id=%s AND worker_number=%s\"\"\",
                    (session.get('sup_id'), wnum))
        if not owns:
            return jerr('You are not assigned to this worker', 403)

    marked_by = session.get('sup_id') or 0
    run(\"\"\"INSERT INTO attendance(worker_number,att_date,status,marked_by)
           VALUES(%s,%s,%s,%s)
           ON CONFLICT (worker_number,att_date)
           DO UPDATE SET status=EXCLUDED.status,
                         marked_by=EXCLUDED.marked_by,
                         marked_at=CURRENT_TIMESTAMP\"\"\",
        (wnum, qdate, status, marked_by))
    return jok(message=f'Saved {wnum} → {status} ({qdate})')


@app.route('/api/attendance/daily', methods=['GET'])
@need_sup
def daily_report():
    err = db_check()
    if err: return err
    qdate = (request.args.get('date') or date.today().isoformat()).strip()
    ws = workers_for_session()
    if not ws:
        return jsonify({'date': qdate, 'summary': {'P': 0, 'A': 0, 'H': 0, 'L': 0, 'U': 0}, 'rows': []})
    nums = [w['number'] for w in ws]
    placeholders = ','.join(['%s'] * len(nums))
    recs = qall(f\"\"\"SELECT worker_number,status FROM attendance
                    WHERE att_date=%s AND worker_number IN ({placeholders})\"\"\",
                tuple([qdate] + nums))
    status_map = {r['worker_number']: r['status'] for r in recs}
    rows = []
    summary = {'P': 0, 'A': 0, 'H': 0, 'L': 0, 'U': 0}
    for w in ws:
        st = status_map.get(w['number'], '')
        rows.append({'number': w['number'], 'name': w['name'], 'status': st or 'U'})
        summary[st if st else 'U'] += 1
    return jsonify({'date': qdate, 'summary': summary, 'rows': rows})


@app.route('/api/attendance/monthly', methods=['GET'])
@need_sup
def monthly_report():
    err = db_check()
    if err: return err
    try:
        year = int(request.args.get('year') or date.today().year)
        month = int(request.args.get('month') or date.today().month)
    except:
        return jerr('Invalid year/month')

    ws = workers_for_session()
    if not ws:
        return jsonify({'year': year, 'month': month, 'rows': []})

    nums = [w['number'] for w in ws]
    placeholders = ','.join(['%s'] * len(nums))
    start = date(year, month, 1).isoformat()
    last_day = calendar.monthrange(year, month)[1]
    end = date(year, month, last_day).isoformat()

    recs = qall(f\"\"\"SELECT worker_number,status,COUNT(*) AS n FROM attendance
                    WHERE worker_number IN ({placeholders})
                      AND att_date BETWEEN %s AND %s
                    GROUP BY worker_number,status\"\"\",
                tuple(nums + [start, end]))

    agg = {}
    for r in recs:
        d2 = agg.setdefault(r['worker_number'], {'P': 0, 'A': 0, 'H': 0, 'L': 0})
        d2[r['status']] = int(r['n'])

    rows = []
    for w in ws:
        a = agg.get(w['number'], {'P': 0, 'A': 0, 'H': 0, 'L': 0})
        present_days = a['P'] + 0.5 * a['H']
        rows.append({
            'number': w['number'],
            'name': w['name'],
            'present': a['P'],
            'absent': a['A'],
            'half': a['H'],
            'leave': a['L'],
            'total_present': present_days
        })
    return jsonify({'year': year, 'month': month, 'days': last_day, 'rows': rows})


@app.route('/api/attendance/download', methods=['GET'])
@need_sup
def download_excel():
    err = db_check()
    if err: return err
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jerr('openpyxl not installed on server. Run: pip install openpyxl', 500)

    try:
        year = int(request.args.get('year') or date.today().year)
        month = int(request.args.get('month') or date.today().month)
    except:
        return jerr('Invalid year/month')

    ws_list = workers_for_session()
    nums = [w['number'] for w in ws_list]

    agg = {}
    if nums:
        placeholders = ','.join(['%s'] * len(nums))
        start = date(year, month, 1).isoformat()
        last_day = calendar.monthrange(year, month)[1]
        end = date(year, month, last_day).isoformat()
        recs = qall(f\"\"\"SELECT worker_number,status,COUNT(*) AS n FROM attendance
                        WHERE worker_number IN ({placeholders})
                          AND att_date BETWEEN %s AND %s
                        GROUP BY worker_number,status\"\"\",
                    tuple(nums + [start, end]))
        for r in recs:
            d2 = agg.setdefault(r['worker_number'], {'P': 0, 'A': 0, 'H': 0, 'L': 0})
            d2[r['status']] = int(r['n'])

    wb = openpyxl.Workbook()
    sh = wb.active
    month_name = calendar.month_name[month]
    sh.title = f\"{month_name} {year}\"

    # Title
    sh.merge_cells('A1:D1')
    sh['A1'] = f\"Attendance Report — {month_name} {year}\"
    sh['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    sh['A1'].fill = PatternFill('solid', fgColor='3949AB')
    sh['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sh.row_dimensions[1].height = 26

    # Header
    headers = ['Worker Number', 'Worker Name', 'Total Present Days', 'Status Breakdown']
    for i, h in enumerate(headers, 1):
        c = sh.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='5C6BC0')
        c.alignment = Alignment(horizontal='center')

    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 4
    for w in ws_list:
        a = agg.get(w['number'], {'P': 0, 'A': 0, 'H': 0, 'L': 0})
        total = a['P'] + 0.5 * a['H']
        sh.cell(row=r, column=1, value=w['number']).border = border
        sh.cell(row=r, column=2, value=w['name']).border = border
        sh.cell(row=r, column=3, value=total).border = border
        sh.cell(row=r, column=4,
                value=f\"P:{a['P']}  A:{a['A']}  H:{a['H']}  L:{a['L']}\").border = border
        r += 1

    if not ws_list:
        sh.cell(row=4, column=1, value=\"No workers assigned.\")

    sh.column_dimensions['A'].width = 16
    sh.column_dimensions['B'].width = 28
    sh.column_dimensions['C'].width = 22
    sh.column_dimensions['D'].width = 32

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f\"Attendance_{month_name}_{year}.xlsx\"
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


# ══════════════════════════════════════════════════════════
# MANAGER — supervisor list + worker assignment
# ══════════════════════════════════════════════════════════
@app.route('/api/manager/supervisors', methods=['GET'])
@need_mgr
def mgr_list_sups():
    err = db_check()
    if err: return err
    sups = qall(\"SELECT id,username,created FROM supervisor ORDER BY username\")
    for s in sups:
        s['workers'] = qall(\"\"\"SELECT w.number,w.name FROM supervisor_workers sw
                               JOIN workers w ON w.number=sw.worker_number
                               WHERE sw.supervisor_id=%s ORDER BY w.number\"\"\", (s['id'],))
    return jsonify(sups)


@app.route('/api/manager/unassigned-workers', methods=['GET'])
@need_mgr
def mgr_unassigned():
    err = db_check()
    if err: return err
    rows = qall(\"\"\"SELECT w.* FROM workers w
                   WHERE NOT EXISTS (SELECT 1 FROM supervisor_workers sw
                                     WHERE sw.worker_number=w.number)
                   ORDER BY w.number\"\"\")
    return jsonify(rows)


@app.route('/api/manager/assign', methods=['POST'])
@need_mgr
def mgr_assign():
    err = db_check()
    if err: return err
    d = jdata()
    try:
        sid = int(d.get('supervisor_id'))
    except:
        return jerr('supervisor_id required')
    wnum = str(d.get('worker_number', '') or '').strip()
    if not wnum:
        return jerr('worker_number required')
    if not qone(\"SELECT id FROM supervisor WHERE id=%s\", (sid,)):
        return jerr('Supervisor not found')
    if not qone(\"SELECT number FROM workers WHERE number=%s\", (wnum,)):
        return jerr('Worker not found')
    # Re-assign (unique on worker_number => move it)
    run(\"\"\"INSERT INTO supervisor_workers(supervisor_id,worker_number)
           VALUES(%s,%s)
           ON CONFLICT (worker_number)
           DO UPDATE SET supervisor_id=EXCLUDED.supervisor_id\"\"\", (sid, wnum))
    return jok(message=f'Worker {wnum} assigned')


@app.route('/api/manager/unassign', methods=['POST'])
@need_mgr
def mgr_unassign():
    err = db_check()
    if err: return err
    d = jdata()
    wnum = str(d.get('worker_number', '') or '').strip()
    if not wnum: return jerr('worker_number required')
    run(\"DELETE FROM supervisor_workers WHERE worker_number=%s\", (wnum,))
    return jok(message=f'Worker {wnum} unassigned')


@app.route('/api/manager/supervisor/<int:sid>', methods=['DELETE'])
@need_mgr
def mgr_del_sup(sid):
    err = db_check()
    if err: return err
    run(\"DELETE FROM supervisor WHERE id=%s\", (sid,))
    return jok(message='Supervisor removed')


# ══════════════════════════════════════════════════════════
if __name__ == '__main__':
    print(\"\n\" + \"=\" * 52)
    print(\"  Billing System Pro\")
    print(\"  Open: http://localhost:5000\")
    print(\"  Supervisor: admin / admin123\")
    print(\"  Manager:    manager / manager123\")
    print(\"=\" * 52 + \"\n\")
    app.run(host='0.0.0.0', port=5000, debug=False)
"
Observation: Create successful: /app/output/app.py
